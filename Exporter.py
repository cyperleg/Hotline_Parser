import openpyxl
from openpyxl import Workbook


HEADER = ["Brand", "Model", "Price", "Description", "Specification", "IMG", "Link"]


class Exporter:
    def __init__(self, name):
        try:
            self.__wb = openpyxl.load_workbook(name + ".xlsx")
            self.__flag = False
        except Exception as e:
            self.__wb = Workbook()
            self.__flag = True
        self.__ws = self.__wb.active
        self.name = name
        self.__url_lst = list()

    def create_sheet(self, name):
        if self.__flag:
            self.__ws = self.__wb.create_sheet(name)
            self.__ws.append(HEADER)
        else:
            try:
                self.__ws = self.__wb.get_sheet_by_name(name)
                self.__get_lst_url()
            except:
                self.__flag = True
                self.__url_lst = list()
                self.create_sheet(name)

    def add_line(self, lst):
        row = lst[0:-2] + [';'.join(lst[-2])] + [lst[-1]]
        self.__ws.append(row)

    def save(self):
        self.__wb.save(f"{self.name}.xlsx")

    def __get_lst_url(self):
        for i, row in enumerate(self.__ws):
            if i == 0:
                continue
            self.__url_lst.append(row[-1].value)

    def check_item(self, url):
        return url in self.__url_lst





