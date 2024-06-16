from selenium import webdriver
from PIL import Image
from io import BytesIO
from selenium.webdriver.common.by import By
import base64
import openpyxl

driver = webdriver.Chrome()

wb = openpyxl.load_workbook("out.xlsx")

sheets = wb.sheetnames

for i in sheets:
    ws = wb[i]
    for k in range(2, ws.max_row + 1):
        imgs = ws.cell(row=k, column=6).value.split(";")
        print(imgs)
        for x, link in enumerate(imgs):
            driver.get(link)
            img = Image.open(
                BytesIO(base64.b64decode(driver.find_element(By.XPATH, "/html/body/img").screenshot_as_base64)))
            ws.add_image(img, )

